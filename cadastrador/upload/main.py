from flask import Flask, jsonify, request
import psycopg2
from psycopg2 import sql
import csv
from decouple import config
from datetime import datetime
import boto3
from botocore.exceptions import NoCredentialsError
import os
import xlrd
import openpyxl
from werkzeug.utils import secure_filename
import pandas as pd

app = Flask(__name__)

UPLOAD_FOLDER = r'C:\Users\980184\Desktop\QQtech\BeaBa\administrador\upload\uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

DATABASE_HOST = config('host')
DATABASE_NAME = config('db')
DATABASE_USER = config('user')
DATABASE_PASSWORD = config('passwd')

AWS_ACCESS_KEY_ID = config('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = config('AWS_SECRET_ACCESS_KEY')
AWS_REGION = config('AWS_REGION')
AWS_BUCKET_NAME = config('AWS_BUCKET_NAME')

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv', 'xls', 'xlsx'}

def obter_informacoes_colunas_banco(host, database, user, password, tabela, referencia_template):
    tipos_colunas = {}
    try:
        conn = psycopg2.connect(
            host=host,
            database=database,
            user=user,
            password=password
        )
        cursor = conn.cursor()

        cursor.execute(
            sql.SQL("SELECT nome_campo, tipo_dado_campo FROM \"BeaBa\".campos WHERE referencia_template = %s"),
            (referencia_template,)
        )

        rows = cursor.fetchall()
        for row in rows:
            coluna = row[0]
            tipo = row[1]
            tipos_colunas[coluna] = tipo
        conn.close()
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {str(e)}")
        return None
    return tipos_colunas

def determinar_tipo(valor):
    if valor.strip() == '':
        return 'undefined'
    try:
        int(valor)
        return 'int'
    except ValueError:
        try:
            float(valor)
            return 'numeric'
        except ValueError:
            if valor.lower() == 'true' or valor.lower() == 'false' or valor.lower() == 'falso' or valor.lower() == 'verdadeiro':
                return 'boolean'
            else:
                try:
                    datetime.strptime(valor, '%Y-%m-%d')
                    return 'date'
                except ValueError:
                    try:
                        datetime.strptime(valor, '%Y-%m-%d %H:%M:%S')
                        return 'timestamp'
                    except ValueError:
                        return 'varchar'

def verificar_tipos_e_nomes_csv(caminho_arquivo, tipos_colunas):
    erros_colunas = []
    erros_tipos = {}

    try:
        with open(caminho_arquivo, newline='') as arquivo_csv:
            leitor = csv.reader(arquivo_csv)
            cabecalho = next(leitor)

            for coluna_banco in tipos_colunas.keys():
                if coluna_banco not in cabecalho:
                    erros_colunas.append(f"Coluna '{coluna_banco}' do template não encontrada no CSV.")

            if erros_colunas:
                return erros_colunas, erros_tipos

            for coluna in cabecalho:
                erros_tipos[coluna] = []

            for i, linha in enumerate(leitor, start=2):
                linhas=[]
                for j, valor in enumerate(linha):
                    coluna = cabecalho[j]
                    tipo_valor = determinar_tipo(valor)
                    tipo_esperado = tipos_colunas.get(coluna)

                    if tipo_esperado is not None and tipo_valor != tipo_esperado:
                        erros_tipos[coluna].append(f"{i}")

    except FileNotFoundError:
        print(f"O arquivo {caminho_arquivo} não foi encontrado.")
        return None, None
    except Exception as e:
        print(f"Ocorreu um erro ao ler o arquivo CSV: {str(e)}")
        return None, None

    erros_tipos = {coluna: erros for coluna, erros in erros_tipos.items() if erros}

    return erros_colunas, erros_tipos

def verificar_tipos_e_nomes_xls(caminho_arquivo, tipos_colunas):
    erros_colunas = []
    erros_tipos = {}

    try:
        workbook = xlrd.open_workbook(caminho_arquivo)
        sheet = workbook.sheet_by_index(0)  

        cabecalho = [str(sheet.cell_value(0, i)).strip() for i in range(sheet.ncols)]

        for coluna_banco in tipos_colunas.keys():
            if coluna_banco not in cabecalho:
                erros_colunas.append(f"Coluna '{coluna_banco}' do template não encontrada no XLS.")

        if erros_colunas:
            return erros_colunas, erros_tipos

        for coluna in cabecalho:
            erros_tipos[coluna] = []

        for i in range(1, sheet.nrows):  
            for j in range(sheet.ncols):
                coluna = cabecalho[j]
                valor = sheet.cell_value(i, j)
                tipo_valor = determinar_tipo(str(valor))
                tipo_esperado = tipos_colunas.get(coluna)

                if tipo_esperado is not None and tipo_valor != tipo_esperado:
                    erros_tipos[coluna].append(f"{i + 1}")

    except xlrd.XLRDError as e:
        print(f"Erro ao ler o arquivo XLS: {str(e)}")
        return [f"Erro ao ler o arquivo XLS: {str(e)}"], None
    except FileNotFoundError:
        print(f"O arquivo {caminho_arquivo} não foi encontrado.")
        return None, None
    except Exception as e:
        print(f"Ocorreu um erro ao ler o arquivo XLS: {str(e)}")
        return None, None

    erros_tipos = {coluna: erros for coluna, erros in erros_tipos.items() if erros}

    return erros_colunas, erros_tipos

def verificar_tipos_e_nomes_xlsx(caminho_arquivo, tipos_colunas):
    erros_colunas = []
    erros_tipos = {}

    try:
        workbook = openpyxl.load_workbook(caminho_arquivo)
        sheet = workbook.active

        cabecalho = [str(sheet.cell(1, i).value).strip() for i in range(1, sheet.max_column + 1)]

        for coluna_banco in tipos_colunas.keys():
            if coluna_banco not in cabecalho:
                erros_colunas.append(f"Coluna '{coluna_banco}' do template não encontrada no XLSX.")

        if erros_colunas:
            return erros_colunas, erros_tipos

        for coluna in cabecalho:
            erros_tipos[coluna] = []

        for row in sheet.iter_rows(min_row=2):
            for j, cell in enumerate(row):
                coluna = cabecalho[j]
                valor = cell.value
                tipo_valor = determinar_tipo(str(valor))
                tipo_esperado = tipos_colunas.get(coluna)

                if tipo_esperado is not None and tipo_valor != tipo_esperado:
                    erros_tipos[coluna].append(f"{cell.row}")

    except FileNotFoundError:
        print(f"O arquivo {caminho_arquivo} não foi encontrado.")
        return None, None
    except Exception as e:
        print(f"Ocorreu um erro ao ler o arquivo XLSX: {str(e)}")
        return None, None

    erros_tipos = {coluna: erros for coluna, erros in erros_tipos.items() if erros}

    return erros_colunas, erros_tipos

def enviar_arquivo_aws(caminho_arquivo, nome_arquivo):
    s3 = boto3.client(
        's3',
        aws_access_key_id=AWS_ACCESS_KEY_ID,
        aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
        region_name=AWS_REGION
    )
    diretorio = request.form.get('diretorio')
    area = request.form.get('area')
    squad = request.form.get('squad')
    pasta_s3 = f'arquivos/{area}/{squad}/{diretorio}/'

    try:
        s3.upload_file(caminho_arquivo, AWS_BUCKET_NAME, pasta_s3 + nome_arquivo)
        url_arquivo = f"https://{AWS_BUCKET_NAME}.s3.amazonaws.com/{pasta_s3}{nome_arquivo}"
        return url_arquivo
    except NoCredentialsError:
        print("Credenciais da AWS não configuradas")
        return None

@app.route('/administrador/upload/templates/index.html', methods=['GET'])
def get_template():
    try:
        with open(r'C:\Users\980184\Desktop\QQtech\BeaBa\administrador\upload\templates\index.html', 'r') as file:
            content = file.read()
        return content
    except FileNotFoundError:
        return "Arquivo não encontrado", 404

@app.route('/upload_file/<int:referencia_template>', methods=['POST'])
def upload_file(referencia_template):
    if 'file' not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"})

    file = request.files['file']

    if file.filename == '':
        return jsonify({"error": "Nome do arquivo vazio"}), 400

    if file and allowed_file(file.filename):
        try:
            tipos_colunas = obter_informacoes_colunas_banco(DATABASE_HOST, DATABASE_NAME, DATABASE_USER, DATABASE_PASSWORD, "BeaBa.campos", referencia_template)

            if tipos_colunas:
                if not os.path.exists(app.config['UPLOAD_FOLDER']):
                    os.makedirs(app.config['UPLOAD_FOLDER'])

                filename = secure_filename(file.filename)
                file_extension = filename.rsplit('.', 1)[1].lower()


                caminho_arquivo = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                if file_extension == 'csv':
                    file.save(caminho_arquivo)
                    erros_colunas, erros_tipos = verificar_tipos_e_nomes_csv(caminho_arquivo, tipos_colunas)
                elif file_extension in ['xls']:
                    file.save(caminho_arquivo)
                    erros_colunas, erros_tipos = verificar_tipos_e_nomes_xls(caminho_arquivo, tipos_colunas)
                elif file_extension in ['xlsx']:
                    file.save(caminho_arquivo)
                    erros_colunas, erros_tipos = verificar_tipos_e_nomes_xlsx(caminho_arquivo, tipos_colunas)
                else:
                    return jsonify({"error": "Formato de arquivo não suportado"}), 400

                if erros_colunas:
                    erros_colunas_detalhados = "\n".join(erros_colunas)
                    return jsonify({"error": f"Erro nas colunas do arquivo:\n{erros_colunas_detalhados}"}), 400
                elif erros_tipos:
                    erros_tipos_detalhados = "\n".join([f"Erros nas linhas da coluna '{coluna}':\n{', '.join(erros)}" for coluna, erros in erros_tipos.items()])
                    return jsonify({"error": f"Erro nos tipos de dados do arquivo:\n{erros_tipos_detalhados}"}), 400
                else:
                    print("Verificação bem-sucedida")
                    
                    timestamp = datetime.now().strftime('%Y-%m-%d-%H-%M-%S')
                    nome_arquivo_com_timestamp = f"{timestamp}_{filename}"
                    
                    url_arquivo = enviar_arquivo_aws(caminho_arquivo, nome_arquivo_com_timestamp)
                    if url_arquivo:
                        conn = psycopg2.connect(
                            host=DATABASE_HOST,
                            database=DATABASE_NAME,
                            user=DATABASE_USER,
                            password=DATABASE_PASSWORD
                        )
                        cursor = conn.cursor()

                        referencia_usuario = request.form.get('referencia_usuario')
                        referencia_usuario_int = int(referencia_usuario)
                        referencia_squad = request.form.get('squad')
                        diretorio = request.form.get('diretorio')
                        area = request.form.get('area')
                        nomeUsuarioUpload = request.form.get('nomeUsuarioUpload')

                        cursor.execute(
                            sql.SQL("INSERT INTO \"BeaBa\".uploads (data_criacao_upload, nome_upload, caminho_upload, referencia_template, referencia_usuario,referencia_nome,referencia_squad,referencia_area,diretorio) VALUES (%s, %s, %s,%s, %s, %s,%s,%s,%s)"),
                            (datetime.now(), nome_arquivo_com_timestamp, url_arquivo, referencia_template, referencia_usuario_int,nomeUsuarioUpload, referencia_squad ,area, diretorio)
                        )
                        conn.commit()
                        cursor.close()
                        
                        os.remove(caminho_arquivo)
                        
                        return jsonify({"message": "Arquivo enviado e inserido na tabela 'uploads' com sucesso"}), 200
                    else:
                        print("Falha ao enviar arquivo para a AWS S3")
                        return jsonify({"error": "Falha ao enviar arquivo para a AWS S3"}), 500

            else:
                print("Não foi possível obter as informações das colunas do banco de dados.")
                return jsonify({"error": "Não foi possível obter informações das colunas do banco de dados"}), 500
        except NoCredentialsError:
            return jsonify({"error": "Credenciais da AWS não configuradas"})
        except Exception as e:
            os.remove(caminho_arquivo)
            return jsonify({"error": str(e)})

    return jsonify({"error": "Tipo de arquivo não permitido"})
if __name__ == '__main__':
    app.run(port=5500)